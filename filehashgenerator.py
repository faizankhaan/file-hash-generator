"""
File Hash Generator
Author: Muhammad Faizan Khan
Date: October 11, 2023
"""

import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import os
import hashlib
import pandas as pd
import fpdf
from openpyxl import load_workbook
from fpdf import FPDF

class HashGeneratorApp:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("File Hash Generator")

        # Variables
        self.target_dir = tk.StringVar()
        self.selected_hashes = {
            "MD5": tk.BooleanVar(),
            "SHA1": tk.BooleanVar(),
            "SHA256": tk.BooleanVar(),
        }

        # Create a frame for grouping components
        frame = ttk.Frame(self.root)
        frame.grid(row=0, column=0, padx=20, pady=20, sticky="w")

        # GUI Components
        ttk.Label(frame, text="Select Target Directory").grid(row=0, column=0, pady=10, sticky="w")
        ttk.Entry(frame, textvariable=self.target_dir, width=50, state="readonly").grid(row=1, column=0, pady=2, sticky="w")
        ttk.Button(frame, text="Browse", command=self.browse_directory).grid(row=2, column=0, pady=5, sticky="w")

        ttk.Label(frame, text="Select Hashing Algorithm(s)").grid(row=3, column=0, pady=10, sticky="w")
        for i, algorithm in enumerate(self.selected_hashes.keys()):
            ttk.Checkbutton(frame, text=algorithm, variable=self.selected_hashes[algorithm]).grid(row=4+i, column=0, pady=5, sticky="w")

        ttk.Button(frame, text="Generate Hashes", command=self.generate_hashes).grid(row=9, column=0, pady=20)

        # Configure the main window
        self.root.geometry("350x320")
        self.root.resizable(False, False)

    def browse_directory(self):
        folder_selected = filedialog.askdirectory()
        self.target_dir.set(folder_selected)

    def calculate_hash(self, file_path, algorithms):
        hash_values = {}
        for algorithm in algorithms:
            hash_obj = hashlib.new(algorithm)
            with open(file_path, 'rb') as f:
                while chunk := f.read(8192):
                    hash_obj.update(chunk)
            hash_values[algorithm] = hash_obj.hexdigest()
        return hash_values

    def generate_hashes(self):
        target_directory = self.target_dir.get()
        selected_algorithms = [algorithm for algorithm, selected in self.selected_hashes.items() if selected.get()]
        output_data = []

        for root_dir, _, files in os.walk(target_directory):
            for file_name in files:
                file_path = os.path.join(root_dir, file_name)
                relative_path = os.path.relpath(file_path, target_directory)  # Get the relative path
                hash_values = self.calculate_hash(file_path, selected_algorithms)
                row = {
                    "S. No": len(output_data) + 1,
                    "Filepath": relative_path,  # Use the relative path
                }
                row.update({algorithm: hash_values[algorithm] for algorithm in selected_algorithms})
                output_data.append(row)

        #Save as Excel and PDF in the target directory
        output_excel_file = os.path.join(target_directory, "Files Hash List.xlsx")
        self.save_to_excel(output_data, output_excel_file)
        messagebox.showinfo("Info", "Hashes saved as Excel and PDF in the target directory.")

    def save_to_excel(self, data, excel_file):
        df = pd.DataFrame(data)
        df.to_excel(excel_file, index=False)

        # Convert Excel to PDF and fit columns on a single page
        pdf_file = excel_file.replace(".xlsx", ".pdf")
        self.convert_excel_to_pdf(excel_file, pdf_file)

    def convert_excel_to_pdf(self, excel_file, pdf_file):
        wb = load_workbook(excel_file)
        ws = wb.active

        pdf = FPDF()
        pdf.add_page()
        pdf.set_auto_page_break(auto=True, margin=10)  # Reduce margin to fit text
        pdf.set_font("Arial", size=3)  # Reduce font size

        col_widths = []
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 4  # Adjust width and font size for PDF
            col_widths.append(adjusted_width)

        total_width = sum(col_widths)
        scale = 190 / total_width  # Adjust scaling factor to fit text better

        for width in col_widths:
            pdf.cell(width * scale, 7, "", border=1)
        pdf.ln()

        for row in ws.iter_rows():
            for i, cell in enumerate(row):
                pdf.cell(col_widths[i] * scale, 2, str(cell.value), border=1)
            pdf.ln()

        pdf.output(pdf_file)


if __name__ == "__main__":
    app = HashGeneratorApp()
    app.root.mainloop()
